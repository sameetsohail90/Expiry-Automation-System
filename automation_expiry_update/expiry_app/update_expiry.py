"""
devices/update_expiry.py
-------------------------
Excel se links read karta hai, login karta hai, har link ki expiry field
mein (current date + 1 month) date daal kar save karta hai, result wapis
Excel mein likh deta hai.

Apni site ke mutabiq neeche CONFIG section change karein.
"""

import time
from datetime import datetime
from dateutil.relativedelta import relativedelta

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service


# ============== CONFIG (apni site ke mutabiq change karein) ==============
LOGIN_URL = "https://yoursite/index.php?r=site/login"
USERNAME = "admin@devaj.co"
PASSWORD = "123"

LOGIN_USERNAME_SELECTOR = (By.NAME, "LoginForm[username]")
LOGIN_PASSWORD_SELECTOR = (By.NAME, "LoginForm[password]")
LOGIN_SUBMIT_SELECTOR = (By.CSS_SELECTOR, "button[type='submit']")

EXPIRY_FIELD_SELECTOR = (By.NAME, "Device[expiry_date]")   # actual field name yahan dalein
SAVE_BUTTON_SELECTOR = (By.CSS_SELECTOR, "button[type='submit']")

DATE_FORMAT = "%Y-%m-%d %H:%M:%S"
# Field "Date and Time" leta hai is liye format mein time bhi shamil hai.
# Agar field ka format alag hai to neeche se uncomment/match karein:
#   "%Y-%m-%d"            -> 2026-07-30
#   "%d-%m-%Y"             -> 30-07-2026
#   "%Y-%m-%dT%H:%M"        -> 2026-07-30T14:30   (HTML datetime-local input)
#   "%m/%d/%Y %I:%M %p"     -> 07/30/2026 02:30 PM

HEADLESS = False           # Pehli baar False rakhein taake browser dikhe aur masla samajh aaye
WAIT_SECONDS = 15
# ===========================================================================


def get_driver():
    options = Options()
    if HEADLESS:
        options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1366,900")
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)


def click_element(driver, element):
    """Click karta hai, normal click fail ho to JS click force karta hai."""
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
    time.sleep(0.5)
    try:
        element.click()
    except Exception as e:
        print(f"  -> Normal click fail hua ({e}), JS click try kar raha hoon...")
        driver.execute_script("arguments[0].click();", element)


def find_submit_button(driver):
    """Save/Submit button dhoondne ki koshish multiple tareeqon se, sabse specific pehle."""
    candidates = [
        # Pehle "Save"/"Update" likhe buttons (sabse reliable, neeche ka koi unrelated button nahi pakrega)
        (By.XPATH, "//button[contains(translate(text(),'SAVE','save'),'save')]"),
        (By.XPATH, "//button[contains(translate(text(),'UPDATE','update'),'update')]"),
        (By.XPATH, "//input[@value[contains(translate(.,'SAVE','save'),'save')]]"),
        (By.XPATH, "//input[@value[contains(translate(.,'UPDATE','update'),'update')]]"),
        # Phir form ke andar wala submit button
        (By.CSS_SELECTOR, "form button[type='submit']"),
        (By.CSS_SELECTOR, "form input[type='submit']"),
        (By.CSS_SELECTOR, "button[type='submit']"),
        (By.CSS_SELECTOR, "input[type='submit']"),
        (By.XPATH, "//button[contains(translate(text(),'LOGIN','login'),'login')]"),
        (By.XPATH, "//input[@value[contains(translate(.,'LOGIN','login'),'login')]]"),
        (By.CSS_SELECTOR, "form button"),
        (By.CSS_SELECTOR, "form input[type='image']"),
    ]
    for by, sel in candidates:
        elems = driver.find_elements(by, sel)
        # Sirf visible elements lo
        visible = [e for e in elems if e.is_displayed()]
        if visible:
            print(f"  -> Button mila selector se: {by} = '{sel}' (text/value: "
                  f"{visible[0].text or visible[0].get_attribute('value')})")
            return visible[0]
    # Kuch na mila to debug ke liye page source aur screenshot save karo
    with open("debug_page_source.html", "w", encoding="utf-8") as f:
        f.write(driver.page_source)
    driver.save_screenshot("debug_screenshot.png")
    raise Exception(
        "Submit button nahi mila. debug_page_source.html aur debug_screenshot.png "
        "check karein taake actual button selector pata chal sake."
    )


def discover_page_elements(driver, label=""):
    """
    Page par mojood saare input/select/textarea aur button elements list karta hai.
    Isse selector manually dhoondhne ki zaroorat nahi - console output se hi pata
    chal jata hai ke expiry field aur save button ka asal name/id kya hai.
    """
    print(f"\n{'='*70}")
    print(f"PAGE ELEMENTS DISCOVERY {('- ' + label) if label else ''}")
    print(f"URL: {driver.current_url}")
    print(f"{'='*70}")

    inputs = driver.find_elements(By.TAG_NAME, "input")
    selects = driver.find_elements(By.TAG_NAME, "select")
    textareas = driver.find_elements(By.TAG_NAME, "textarea")
    buttons = driver.find_elements(By.TAG_NAME, "button")

    print(f"\n--- INPUT fields ({len(inputs)}) ---")
    for el in inputs:
        if not el.is_displayed():
            continue
        print(f"  type={el.get_attribute('type'):10s} "
              f"name={str(el.get_attribute('name')):30s} "
              f"id={str(el.get_attribute('id')):25s} "
              f"value={el.get_attribute('value')}")

    print(f"\n--- SELECT fields ({len(selects)}) ---")
    for el in selects:
        if not el.is_displayed():
            continue
        print(f"  name={str(el.get_attribute('name')):30s} id={el.get_attribute('id')}")

    print(f"\n--- TEXTAREA fields ({len(textareas)}) ---")
    for el in textareas:
        if not el.is_displayed():
            continue
        print(f"  name={str(el.get_attribute('name')):30s} id={el.get_attribute('id')}")

    print(f"\n--- BUTTONS ({len(buttons)}) ---")
    for el in buttons:
        if not el.is_displayed():
            continue
        print(f"  type={str(el.get_attribute('type')):10s} "
              f"text='{el.text.strip()}' "
              f"id={el.get_attribute('id')} "
              f"class={el.get_attribute('class')}")

    print(f"{'='*70}\n")

    with open(f"debug_elements_{label or 'page'}.html", "w", encoding="utf-8") as f:
        f.write(driver.page_source)
    driver.save_screenshot(f"debug_elements_{label or 'page'}.png")


def click_save(driver):
    """
    Save button click karta hai. Priority:
    1. Agar page par JS function save() defined hai, usko DIRECTLY call karo (sabse reliable).
    2. Warna <a onclick="save();">, <button>, <input type=submit> dhoondo aur click karo.
    """
    # Tareeqa 1: directly JS function call (yeh aapke <a onclick="save();"> ke liye perfect hai)
    has_save_fn = driver.execute_script(
        "return typeof save === 'function';"
    )
    if has_save_fn:
        print("  -> JS function save() mil gaya, directly call kar raha hoon...")
        driver.execute_script("save();")
        return

    # Tareeqa 2: <a> tag jiska onclick mein save() ho
    anchors = driver.find_elements(By.XPATH, "//a[contains(@onclick,'save')]")
    visible_anchors = [a for a in anchors if a.is_displayed()]
    if visible_anchors:
        print(f"  -> <a> save button mila (onclick='{visible_anchors[0].get_attribute('onclick')}')")
        click_element(driver, visible_anchors[0])
        return

    # Tareeqa 3: <a> tag jiska text "Save" ho
    anchors = driver.find_elements(By.XPATH, "//a[contains(translate(text(),'SAVE','save'),'save')]")
    visible_anchors = [a for a in anchors if a.is_displayed()]
    if visible_anchors:
        print(f"  -> <a> 'Save' text wala button mila")
        click_element(driver, visible_anchors[0])
        return

    # Tareeqa 4: purana button/input based fallback
    click_element(driver, find_submit_button(driver))


def login(driver, login_url=None, username=None, password=None):
    login_url = login_url or LOGIN_URL
    username = username or USERNAME
    password = password or PASSWORD

    driver.get(login_url)
    wait = WebDriverWait(driver, WAIT_SECONDS)
    discover_page_elements(driver, label="login_page")
    wait.until(EC.presence_of_element_located(LOGIN_USERNAME_SELECTOR)).send_keys(username)
    driver.find_element(*LOGIN_PASSWORD_SELECTOR).send_keys(password)
    click_element(driver, find_submit_button(driver))
    time.sleep(2)


def set_field_value(driver, field, value):
    """
    Field mein value set karta hai — pehle normal tareeqa try karta hai,
    phir JS se force set karta hai (readonly / datepicker fields ke liye zaroori).
    """
    # Tareeqa 1: normal clear + type
    try:
        field.click()
        field.send_keys(Keys.CONTROL, "a")
    except Exception:
        pass

    driver.execute_script("""
        var el = arguments[0];
        var val = arguments[1];
        el.removeAttribute('readonly');
        el.value = val;
        el.dispatchEvent(new Event('input', { bubbles: true }));
        el.dispatchEvent(new Event('change', { bubbles: true }));
        el.dispatchEvent(new Event('blur', { bubbles: true }));
    """, field, value)

    # Confirm value actually set hui
    actual_value = field.get_attribute("value")
    return actual_value


def update_single_link(driver, link, manual_expiry=None):
    wait = WebDriverWait(driver, WAIT_SECONDS)
    print(f"  -> Opening: {link}")
    driver.get(link)

    print(f"  -> Looking for expiry field: {EXPIRY_FIELD_SELECTOR}")
    discover_page_elements(driver, label="edit_page")
    try:
        expiry_field = wait.until(EC.presence_of_element_located(EXPIRY_FIELD_SELECTOR))
    except Exception:
        with open("debug_expiry_page.html", "w", encoding="utf-8") as f:
            f.write(driver.page_source)
        driver.save_screenshot("debug_expiry_page.png")
        raise Exception(
            "Expiry field nahi mili is page par. debug_expiry_page.html aur "
            "debug_expiry_page.png check karein — Inspect Element se actual "
            "field ka 'name' attribute confirm karein."
        )

    old_value = expiry_field.get_attribute("value")

    if manual_expiry:
        new_expiry = manual_expiry
    else:
        new_expiry = (datetime.now() + relativedelta(months=1)).strftime(DATE_FORMAT)

    print(f"  -> Old value: {old_value}")
    print(f"  -> Setting expiry to: {new_expiry}")

    actual_value = set_field_value(driver, expiry_field, new_expiry)
    print(f"  -> Field value after set: {actual_value}")

    if actual_value != new_expiry:
        print(f"  -> WARNING: Field value set nahi hui jaisi expect thi! "
              f"Yeh shayad JS datepicker hai jo sirf calendar click accept karta hai.")

    click_save(driver)
    time.sleep(2)
    print(f"  -> Saved (button clicked).")

    return new_expiry


def run_expiry_update(input_path, output_path, manual_expiry=None):
    """Excel file process karke result return karta hai. Django view isko call karega.

    Excel columns:
      - Link        (required) -> device edit page URL
      - Login URL   (optional) -> agar diya ho to isi se login hoga, warna global LOGIN_URL
      - Username    (optional) -> agar diya ho to isi se login hoga, warna global USERNAME
      - Password    (optional) -> agar diya ho to isi se login hoga, warna global PASSWORD

    Same (Login URL, Username, Password) wale rows ek "site group" maan ke
    sirf EK dafa login hota hai, phir us group ke saare links process hote hain.
    Isse alag-alag sites/clients ke links ek hi Excel mein dalna possible hai.

    manual_expiry: agar user ne specific date/time diya ho (e.g. "2026-07-30 14:30:00"),
                   to wahi use hogi. None ho to auto "+1 month from today" calculate hogi.
    """
    wb_in = openpyxl.load_workbook(input_path, data_only=True)
    ws_in = wb_in.active

    headers = [c.value for c in ws_in[1]]
    if "Link" not in headers:
        raise ValueError("Excel mein 'Link' naam ka column hona chahiye (Row 1 header).")
    link_col = headers.index("Link") + 1

    login_url_col = headers.index("Login URL") + 1 if "Login URL" in headers else None
    username_col = headers.index("Username") + 1 if "Username" in headers else None
    password_col = headers.index("Password") + 1 if "Password" in headers else None

    # Sab rows collect karo (link + iska login info)
    rows = []
    for row in range(2, ws_in.max_row + 1):
        link = ws_in.cell(row=row, column=link_col).value
        if not link:
            continue
        link = str(link).strip()

        row_login_url = str(ws_in.cell(row=row, column=login_url_col).value).strip() \
            if login_url_col and ws_in.cell(row=row, column=login_url_col).value else LOGIN_URL
        row_username = str(ws_in.cell(row=row, column=username_col).value).strip() \
            if username_col and ws_in.cell(row=row, column=username_col).value else USERNAME
        row_password = str(ws_in.cell(row=row, column=password_col).value).strip() \
            if password_col and ws_in.cell(row=row, column=password_col).value else PASSWORD

        rows.append({
            "link": link,
            "login_url": row_login_url,
            "username": row_username,
            "password": row_password,
        })

    print(f"Total links found in Excel: {len(rows)}")
    for i, r in enumerate(rows, 1):
        print(f"  {i}. {r['link']}  (login: {r['login_url']})")

    if not rows:
        raise ValueError("Excel ke 'Link' column mein koi bhi link nahi mila.")

    # Same (login_url, username, password) wale rows ko group karo
    groups = {}
    for r in rows:
        key = (r["login_url"], r["username"], r["password"])
        groups.setdefault(key, []).append(r["link"])

    print(f"\nTotal site groups (alag login): {len(groups)}\n")

    # Naya CLEAN output workbook banayenge (merged-cell / read-only issues se bachne ke liye)
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Result"
    ws_out.append(["Link", "Login URL", "Status", "New Expiry"])

    results = []
    driver = get_driver()
    try:
        for group_idx, ((login_url, username, password), links) in enumerate(groups.items(), 1):
            print(f"=== Site Group {group_idx}/{len(groups)}: {login_url} ===")
            print(f"Logging in as {username} ...")
            login(driver, login_url, username, password)
            print("Login done.\n")

            for idx, link in enumerate(links, 1):
                print(f"[{idx}/{len(links)}] Processing: {link}")
                try:
                    new_expiry = update_single_link(driver, link, manual_expiry=manual_expiry)
                    ws_out.append([link, login_url, "Success", new_expiry])
                    results.append((link, "Success", new_expiry))
                    print(f"  [OK] New expiry: {new_expiry}\n")
                except Exception as e:
                    err_msg = f"Failed: {e}"
                    ws_out.append([link, login_url, err_msg, ""])
                    results.append((link, err_msg, ""))
                    print(f"  [FAIL] {e}\n")
    finally:
        driver.quit()
        wb_out.save(output_path)
        print(f"Result saved to: {output_path}")

    return results


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--date", required=False, default=None,
                         help="Manual expiry e.g. '2026-07-30 14:30:00'. Na diya to +1 month auto. "
                              "Excel mein 'Link' (required) + optional 'Login URL'/'Username'/'Password' "
                              "columns ho sakte hain agar alag sites ke links hon.")
    args = parser.parse_args()
    res = run_expiry_update(args.input, args.output, manual_expiry=args.date)
    for link, status, expiry in res:
        print(f"{status:10s} | {expiry:10s} | {link}")